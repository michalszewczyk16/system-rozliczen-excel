import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class HRReportGenerator:
    def __init__(self, pracownicy_file='pracownicy.xlsx', godziny_pattern='godziny_tydzien*.xlsx'):
        self.pracownicy_file = pracownicy_file
        self.godziny_pattern = godziny_pattern

    def generuj(self, wyjscie='raport_koncowy.xlsx'):
        df_prac = pd.read_excel(self.pracownicy_file)
        pliki = glob.glob(self.godziny_pattern)
        df_godziny = pd.concat([pd.read_excel(f) for f in pliki])

        df_godziny = df_godziny[df_godziny['Przepracowane_Godziny'] >= 0]

        df_suma = df_godziny.groupby('ID_Pracownika')['Przepracowane_Godziny'].sum().reset_index()
        df_raport = pd.merge(df_prac, df_suma, on='ID_Pracownika', how='left').fillna(0)
        df_raport['Wynagrodzenie_Calkowite'] = df_raport['Przepracowane_Godziny'] * df_raport['Stawka_Godzinowa']

        with pd.ExcelWriter(wyjscie, engine='openpyxl') as writer:
            df_raport.to_excel(writer, index=False, sheet_name='Szczegóły')

        wb = load_workbook(wyjscie)
        ws = wb['Szczegóły']
        naglowek_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        for cell in ws[1]:
            cell.fill = naglowek_fill
            cell.font = Font(color="FFFFFF", bold=True)
        wb.save(wyjscie)


if __name__ == "__main__":
    HRReportGenerator().generuj()